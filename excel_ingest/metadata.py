from __future__ import annotations

import hashlib
import os
import re
from collections import Counter
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from excel_ingest.structure import FileStructureMetadata, MergedCellInfo


@dataclass
class ColumnMetadata:
    column_index: int                      # 1-based
    column_letter: str
    hierarchical_header: str               # "[Parent].[Child]" or "[Header]"
    db_canonical_bronze_column_name: str                # unique SQL-safe Delta column name
    raw_header_texts: List[Optional[str]]  # one entry per header row
    section_id: int                        # partition by blank separator columns
    is_blank_column: bool
    is_hidden_column: bool
    is_part_of_merge: bool
    merge_span_cols: int                   # 1 if not merged


@dataclass
class FileMetadata:
    file_id: str
    file_name: str
    sheet_name: str
    total_rows: int
    total_cols: int
    data_row_count: int
    num_sections: int
    num_merged_regions: int
    header_signature: str                  # SHA-256 of canonical header string

    def to_dict(self) -> Dict[str, Any]:
        return {
            "file_id": self.file_id,
            "file_name": self.file_name,
            "sheet_name": self.sheet_name,
            "total_rows": self.total_rows,
            "total_cols": self.total_cols,
            "data_row_count": self.data_row_count,
            "num_sections": self.num_sections,
            "num_merged_regions": self.num_merged_regions,
            "header_signature": self.header_signature,
        }


@dataclass
class MetadataExtractionResult:
    file_metadata: FileMetadata
    column_metadata: List[ColumnMetadata]
    messages: List[str] = field(default_factory=list)

    def column_records(self) -> List[Dict[str, Any]]:
        """Flat list of dicts for every column — ready for display(), DataFrame, or export.

        Column names are human-readable. Use this for inspection and ad-hoc analysis.
        Use to_delta_records() when persisting to Delta tables.
        """
        return [
            {
                "file_id":             self.file_metadata.file_id,
                "file_name":           self.file_metadata.file_name,
                "col_index":           c.column_index,
                "col_letter":          c.column_letter,
                "hierarchical_header": c.hierarchical_header,
                "db_canonical_bronze_column_name":  c.db_canonical_bronze_column_name,
                "column_group":        c.section_id,
                "is_blank":            c.is_blank_column,
                "is_hidden":           c.is_hidden_column,
                "is_merged":           c.is_part_of_merge,
                "merge_span":          c.merge_span_cols,
            }
            for c in self.column_metadata
        ]

    def bronze_schema(self) -> Dict[str, int]:
        """Column map for this file: {db_canonical_bronze_column_name: column_index}.

        Non-blank columns only. Use to build the file-specific SELECT when loading
        into a superset bronze table — columns absent from this file are NULL-filled::

            schema = meta.bronze_schema()
            # {"employee_id": 1, "first_name": 2, ...}
        """
        return {
            col.db_canonical_bronze_column_name: col.column_index
            for col in self.column_metadata
            if not col.is_blank_column
        }

    def signature_record(self) -> Dict[str, Any]:
        """Minimal dict for schema-change tracking — file identity + header signature only.

        Designed to be stored in a signatures reference table or config store and
        compared across runs to detect column layout changes::

            # Persist after each ingest
            spark.createDataFrame([result.metadata.signature_record()]) \\
                .write.mode("append").saveAsTable("...excel_schema_signatures")

            # Detect drift: if the signature changed, re-map
            if new_sig != stored_sig:
                # schema changed — trigger fresh mapping + alert
        """
        fm = self.file_metadata
        return {
            "file_id":          fm.file_id,
            "file_name":        fm.file_name,
            "sheet_name":       fm.sheet_name,
            "total_cols":       fm.total_cols,
            "header_signature": fm.header_signature,
        }

    def to_delta_records(self) -> List[Dict[str, Any]]:
        return [
            {
                "file_id":             self.file_metadata.file_id,
                "column_index":        c.column_index,
                "column_letter":       c.column_letter,
                "hierarchical_header": c.hierarchical_header,
                "db_canonical_bronze_column_name":  c.db_canonical_bronze_column_name,
                "section_id":          c.section_id,
                "is_blank_column":     c.is_blank_column,
                "is_hidden_column":    c.is_hidden_column,
                "is_part_of_merge":    c.is_part_of_merge,
                "merge_span_cols":     c.merge_span_cols,
            }
            for c in self.column_metadata
        ]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _build_hierarchical_header(texts: List[Optional[str]], separator: str = ".") -> str:
    parts = [t for t in texts if t is not None and t.strip() != ""]
    if not parts:
        return "[Unnamed]"
    if len(parts) == 1:
        return f"[{parts[0]}]"
    return separator.join(f"[{p}]" for p in parts)


def _assign_section_ids(col_count: int, blank_col_indices: List[int]) -> List[int]:
    blank_set = set(blank_col_indices)
    section_ids: List[int] = []
    current = 1
    for col_idx in range(1, col_count + 1):
        if col_idx in blank_set:
            current += 1
        section_ids.append(current)
    return section_ids


def _merge_span_for_col(col_idx: int, merged_cells: List[MergedCellInfo]) -> int:
    for m in merged_cells:
        if m.min_col <= col_idx <= m.max_col and m.span_cols > 1:
            return m.span_cols
    return 1


def _get_horizontal_merge_value(row_idx: int, col_idx: int, merged_cells: List[MergedCellInfo]) -> Optional[str]:
    """Return the parent label when col_idx sits inside a horizontal merge in row_idx.

    Only horizontal merges (same row, multiple columns) propagate their label to
    sibling columns.  Vertical merges (same column, multiple rows) are excluded to
    avoid duplicating values across header levels.
    """
    for m in merged_cells:
        if (m.min_row == row_idx == m.max_row
                and m.min_col < col_idx <= m.max_col):
            return m.top_left_value
    return None


def _sanitise_level(text: str) -> str:
    """Sanitise one hierarchy level into a SQL-safe identifier fragment.

    Rules applied in order:
      & → and  |  % → pct  |  all other non-alphanumeric → _
      squeeze consecutive _ → single _  |  strip leading/trailing _
      prefix col_ if result starts with a digit
    """
    text = text.lower()
    text = text.replace("&", "and")
    text = text.replace("%", "pct")
    text = re.sub(r"[^a-z0-9]", "_", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    if not text:
        return "col"
    if text[0].isdigit():
        return f"col_{text}"
    return text


def _assign_bronze_names(hier_headers: List[str]) -> List[str]:
    """Generate unique SQL-safe bronze column names from a list of hierarchical headers.

    Strategy:
      1. Use leaf segment only when it is unique across the sheet.
      2. Escalate duplicated leaves to the full path (all levels joined with __).
      3. If full-path duplicates still exist (same header repeated verbatim),
         append _N (1-based occurrence) as a final fallback.

    Hierarchy levels are separated by __ (double underscore) to distinguish them
    from single _ used within a level name. Dots are not used because Delta column
    names with dots require backtick-quoting everywhere in SQL.
    """
    all_parts: List[List[str]] = []
    for header in hier_headers:
        raw_parts = re.findall(r"\[([^\]]+)\]", header)
        sanitised = [_sanitise_level(p) for p in raw_parts] if raw_parts else ["col"]
        all_parts.append(sanitised)

    # Attempt 1: leaf only
    names: List[str] = [p[-1] for p in all_parts]

    # Escalate duplicated leaves to full path
    counts = Counter(names)
    dupes = {n for n, c in counts.items() if c > 1}
    if dupes:
        names = [
            "__".join(p) if names[i] in dupes else names[i]
            for i, p in enumerate(all_parts)
        ]

    # Final fallback: append occurrence index for any remaining collisions
    counts = Counter(names)
    if any(v > 1 for v in counts.values()):
        occurrence: Dict[str, int] = {}
        for i, name in enumerate(names):
            if counts[name] > 1:
                occurrence[name] = occurrence.get(name, 0) + 1
                names[i] = f"{name}_{occurrence[name]}"

    return names


def combine_column_records(results: List[MetadataExtractionResult]) -> List[Dict[str, Any]]:
    """Flatten column_records() from a list of MetadataExtractionResult into one list.

    Typical use — display all files in one scrollable table::

        display(spark.createDataFrame(combine_column_records(all_metadata)))
    """
    return [rec for r in results for rec in r.column_records()]


def build_superset_schema(results: List[MetadataExtractionResult]) -> List[str]:
    """Return sorted list of all distinct db_canonical_bronze_column_names across a list of results.

    Use to derive the superset column set for a consolidated bronze table when ingesting
    multiple files with varying structures (e.g. 60 country HR files where some columns
    are GDPR-suppressed or country-specific). Missing columns in individual files load
    as NULL when the superset schema is used::

        all_cols = build_superset_schema(all_sheet1_metadata)
        ddl = ", ".join(f"`{c}` STRING" for c in all_cols)
        spark.sql(f"CREATE TABLE IF NOT EXISTS bronze.hr_sheet1 ({ddl}, source_file STRING)")

        # Per file: meta.bronze_schema() gives {col_name: col_index} for NULL-safe SELECT
    """
    cols = {
        col.db_canonical_bronze_column_name
        for r in results
        for col in r.column_metadata
        if not col.is_blank_column
    }
    return sorted(cols)


def _generate_signature(column_headers: List[str]) -> str:
    canonical = "|".join(column_headers)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_metadata(
    file_path: str,
    structure: FileStructureMetadata,
    file_id: Optional[str] = None,
) -> MetadataExtractionResult:
    msgs: List[str] = []

    if file_id is None:
        file_id = os.path.basename(file_path)

    file_name = os.path.basename(file_path)
    header_struct = structure.header_structure
    merged = structure.merged_cells
    blank_cols = set(structure.blank_column_indices)
    hidden_cols = set(structure.hidden_column_indices)
    total_cols = structure.total_cols

    section_ids = _assign_section_ids(total_cols, structure.blank_column_indices)
    header_rows = header_struct.header_row_indices if header_struct else []

    # First pass: collect hierarchical headers and per-column data
    intermediates: List[Tuple[int, List[Optional[str]], str, bool, int]] = []
    for col_idx in range(1, total_cols + 1):
        raw_texts: List[Optional[str]] = []
        if header_struct:
            for row_idx in header_rows:
                row_vals = header_struct.raw_headers.get(row_idx, [])
                val = row_vals[col_idx - 1] if col_idx - 1 < len(row_vals) else None
                if not val or not str(val).strip():
                    if col_idx not in blank_cols:
                        val = _get_horizontal_merge_value(row_idx, col_idx, merged)
                raw_texts.append(val if val and str(val).strip() else None)
        else:
            raw_texts = [None]

        hier = _build_hierarchical_header(raw_texts)
        in_merge = any(m.min_col <= col_idx <= m.max_col for m in merged)
        span = _merge_span_for_col(col_idx, merged)
        intermediates.append((col_idx, raw_texts, hier, in_merge, span))

    # Generate bronze column names — needs all headers together for uniqueness resolution
    bronze_names = _assign_bronze_names([x[2] for x in intermediates])

    # Second pass: build ColumnMetadata with db_canonical_bronze_column_name assigned
    columns: List[ColumnMetadata] = []
    for (col_idx, raw_texts, hier, in_merge, span), bronze_name in zip(intermediates, bronze_names):
        columns.append(
            ColumnMetadata(
                column_index=col_idx,
                column_letter=get_column_letter(col_idx),
                hierarchical_header=hier,
                db_canonical_bronze_column_name=bronze_name,
                raw_header_texts=raw_texts,
                section_id=section_ids[col_idx - 1],
                is_blank_column=(col_idx in blank_cols),
                is_hidden_column=(col_idx in hidden_cols),
                is_part_of_merge=in_merge,
                merge_span_cols=span,
            )
        )

    all_headers = [c.hierarchical_header for c in columns]
    signature = _generate_signature(all_headers)
    num_sections = max((c.section_id for c in columns), default=1)

    msgs.append(
        f"Extracted metadata for {total_cols} column(s), "
        f"{num_sections} section(s). Signature: {signature[:12]}..."
    )

    file_meta = FileMetadata(
        file_id=file_id,
        file_name=file_name,
        sheet_name=structure.sheet_name,
        total_rows=structure.total_rows,
        total_cols=total_cols,
        data_row_count=structure.data_row_count,
        num_sections=num_sections,
        num_merged_regions=len(merged),
        header_signature=signature,
    )

    return MetadataExtractionResult(
        file_metadata=file_meta,
        column_metadata=columns,
        messages=msgs,
    )
