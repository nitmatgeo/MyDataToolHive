from __future__ import annotations

import os
from dataclasses import dataclass, field
from enum import Enum
from typing import List, Optional, Tuple

from excel_ingest.utils.paths import FileLocationType, detect_location_type


class ValidationStatus(Enum):
    PASSED  = "PASSED"
    WARNING = "WARNING"
    FAILED  = "FAILED"

    @property
    def description(self) -> str:
        return {
            "PASSED":  "File is valid, readable, and all sheets are accessible.",
            "WARNING": "File is readable but has one or more hidden sheets — review if intentional.",
            "FAILED":  "File could not be validated. Check the errors field for details.",
        }[self.value]


@dataclass
class SheetInfo:
    name: str
    index: int
    is_hidden: bool
    is_active: bool


@dataclass
class FileValidationResult:
    file_path: str
    location_type: FileLocationType
    status: ValidationStatus
    file_exists: bool
    file_size_bytes: Optional[int]
    is_excel_format: bool
    format_type: Optional[str]          # "xlsx", "xlsm", "xls"
    is_readable: bool
    is_password_protected: bool
    sheets: List[SheetInfo]
    active_sheet_name: Optional[str]
    messages: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def total_sheets(self) -> int:
        return len(self.sheets)

    @property
    def visible_sheet_names(self) -> List[str]:
        return [s.name for s in self.sheets if not s.is_hidden]

    @property
    def all_sheet_names(self) -> List[str]:
        return [s.name for s in self.sheets]

    def summary_record(self, label: str = "") -> dict:
        """Flat dict for Spark DataFrame display — one row per file."""
        return {
            "label":                 label,
            "file":                  os.path.basename(self.file_path),
            "status":                self.status.value,
            "status_description":    self.status.description,
            "is_readable":           self.is_readable,
            "is_password_protected": self.is_password_protected,
            "file_exists":           self.file_exists,
            "format_type":           self.format_type or "",
            "file_size_bytes":       self.file_size_bytes or 0,
            "total_sheets":          self.total_sheets,
            "visible_sheets":        str(self.visible_sheet_names),
            "warnings":              str(self.warnings) if self.warnings else "",
            "errors":                str(self.errors)   if self.errors   else "",
        }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _check_file_existence(
    file_path: str,
    location_type: FileLocationType,
) -> Tuple[bool, Optional[int], List[str]]:
    msgs: List[str] = []

    if location_type == FileLocationType.AZURE_STORAGE:
        msgs.append("Azure Storage path — existence check delegated to orchestrator.")
        return True, None, msgs

    if location_type in (FileLocationType.VOLUME, FileLocationType.DBFS):
        try:
            dbutils = _get_dbutils()  # type: ignore[assignment]
            if dbutils is not None:
                ls_result = dbutils.fs.ls(file_path)
                size = ls_result[0].size if ls_result else None
                msgs.append(f"File found via dbutils.fs.ls: {file_path}")
                return True, size, msgs
        except Exception as exc:
            msgs.append(f"dbutils.fs.ls failed ({exc}); falling back to OS path check.")
        # Fallback: DBFS paths are also accessible via /dbfs/ mount
        local_path = file_path.replace("dbfs:/", "/dbfs/") if file_path.startswith("dbfs:/") else file_path
        if os.path.exists(local_path):
            size = os.path.getsize(local_path)
            msgs.append(f"File found via OS path: {local_path}")
            return True, size, msgs
        msgs.append(f"File not found: {file_path}")
        return False, None, msgs

    # LOCAL / WORKSPACE / UNKNOWN
    if os.path.exists(file_path):
        size = os.path.getsize(file_path)
        msgs.append(f"File found: {file_path}")
        return True, size, msgs

    msgs.append(f"File not found: {file_path}")
    return False, None, msgs


def _validate_excel_format(file_path: str) -> Tuple[bool, Optional[str], List[str]]:
    msgs: List[str] = []
    lower = file_path.lower()
    if lower.endswith(".xlsx"):
        return True, "xlsx", msgs
    if lower.endswith(".xlsm"):
        return True, "xlsm", msgs
    if lower.endswith(".xls"):
        msgs.append("Legacy .xls format — openpyxl support may be limited.")
        return True, "xls", msgs
    msgs.append(f"Unsupported file extension: {os.path.splitext(file_path)[1]}")
    return False, None, msgs


def _read_workbook(file_path: str, password: Optional[str]):
    import io
    import openpyxl
    local_path = _resolve_local_path(file_path)
    if password is not None:
        # Try msoffcrypto first (handles file-level AES encryption).
        # Falls through to openpyxl worksheet-level password kwarg when:
        #   - msoffcrypto is not installed (ImportError), or
        #   - file is not AES-encrypted (worksheet protection only — msoffcrypto raises FileFormatError etc.)
        # Re-raises only when msoffcrypto confirms a wrong-password failure.
        try:
            import msoffcrypto
        except ImportError:
            raise ImportError(
                "msoffcrypto-tool is required to open AES-encrypted Excel files. "
                "Run: pip install msoffcrypto-tool"
            )
        try:
            with open(local_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
            decrypted.seek(0)
            return openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
        except Exception as exc:
            msg = str(exc).lower()
            if "password" in msg or "decrypt" in msg:
                raise
            # Not AES-encrypted — fall through to openpyxl worksheet-level kwarg
        return openpyxl.load_workbook(local_path, read_only=True, data_only=True, password=password)
    return openpyxl.load_workbook(local_path, read_only=True, data_only=True)


def _check_password_and_sheets(
    file_path: str,
    password: Optional[str],
) -> Tuple[bool, bool, List[SheetInfo], Optional[str], List[str]]:
    msgs: List[str] = []
    try:
        wb = _read_workbook(file_path, password)
        sheets = []
        active_name = wb.active.title if wb.active else None
        for idx, ws in enumerate(wb.worksheets):
            sheets.append(
                SheetInfo(
                    name=ws.title,
                    index=idx,
                    is_hidden=(ws.sheet_state != "visible"),
                    is_active=(ws.title == active_name),
                )
            )
        wb.close()
        msgs.append(f"Workbook readable — {len(sheets)} sheet(s) found.")
        # Decryption succeeded with a password → file IS password-protected.
        return True, password is not None, sheets, active_name, msgs
    except ImportError as exc:
        msgs.append(str(exc))
        return False, False, [], None, msgs
    except Exception as exc:
        msg = str(exc).lower()
        # "not a zip file" / "bad zip file" = openpyxl trying to open an AES-encrypted
        # xlsx without decryption — the file is password-protected but no password was given.
        if ("password" in msg or "encrypted" in msg or "decrypt" in msg
                or "not a zip file" in msg or "bad zip" in msg):
            if password is not None:
                msgs.append("Password provided but incorrect or unsupported encryption.")
            else:
                msgs.append("File is password-protected — provide a password.")
            return False, True, [], None, msgs
        msgs.append(f"Cannot read workbook: {exc}")
        return False, False, [], None, msgs


def _resolve_local_path(file_path: str) -> str:
    if file_path.startswith("dbfs:/"):
        return file_path.replace("dbfs:/", "/dbfs/", 1)
    return file_path


def _get_dbutils():
    try:
        import IPython
        ip = IPython.get_ipython()
        if ip and "dbutils" in ip.user_ns:
            return ip.user_ns["dbutils"]
    except ImportError:
        pass  # IPython not installed — not in a Databricks/Jupyter environment
    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def validate_excel_file(
    file_path: str,
    password: Optional[str] = None,
) -> FileValidationResult:
    location_type = detect_location_type(file_path)
    all_msgs: List[str] = []
    all_warnings: List[str] = []
    all_errors: List[str] = []

    # Stage A: existence
    exists, size, msgs = _check_file_existence(file_path, location_type)
    all_msgs.extend(msgs)
    if not exists:
        all_errors.append(f"File does not exist: {file_path}")
        return FileValidationResult(
            file_path=file_path, location_type=location_type,
            status=ValidationStatus.FAILED, file_exists=False,
            file_size_bytes=None, is_excel_format=False, format_type=None,
            is_readable=False, is_password_protected=False,
            sheets=[], active_sheet_name=None,
            messages=all_msgs, warnings=all_warnings, errors=all_errors,
        )

    # Stage B: format
    is_excel, fmt, msgs = _validate_excel_format(file_path)
    all_msgs.extend(msgs)
    if not is_excel:
        all_errors.append("File is not a recognised Excel format.")
        return FileValidationResult(
            file_path=file_path, location_type=location_type,
            status=ValidationStatus.FAILED, file_exists=True,
            file_size_bytes=size, is_excel_format=False, format_type=None,
            is_readable=False, is_password_protected=False,
            sheets=[], active_sheet_name=None,
            messages=all_msgs, warnings=all_warnings, errors=all_errors,
        )

    # Stage C: read + sheets
    is_readable, is_protected, sheets, active_sheet, msgs = _check_password_and_sheets(
        file_path, password
    )
    all_msgs.extend(msgs)

    if not is_readable:
        all_errors.extend(msgs)
        return FileValidationResult(
            file_path=file_path, location_type=location_type,
            status=ValidationStatus.FAILED, file_exists=True,
            file_size_bytes=size, is_excel_format=True, format_type=fmt,
            is_readable=False, is_password_protected=is_protected,
            sheets=[], active_sheet_name=None,
            messages=all_msgs, warnings=all_warnings, errors=all_errors,
        )

    hidden = [s for s in sheets if s.is_hidden]
    if hidden:
        all_warnings.append(f"{len(hidden)} hidden sheet(s): {[s.name for s in hidden]}")

    status = ValidationStatus.WARNING if all_warnings else ValidationStatus.PASSED

    return FileValidationResult(
        file_path=file_path, location_type=location_type,
        status=status, file_exists=True,
        file_size_bytes=size, is_excel_format=True, format_type=fmt,
        is_readable=True, is_password_protected=is_protected,
        sheets=sheets, active_sheet_name=active_sheet,
        messages=all_msgs, warnings=all_warnings, errors=all_errors,
    )


# ---------------------------------------------------------------------------
# Convenience helpers
# ---------------------------------------------------------------------------

def check_file_exists(file_path: str) -> bool:
    return validate_excel_file(file_path).file_exists


def get_sheet_names(file_path: str, password: Optional[str] = None) -> List[str]:
    return validate_excel_file(file_path, password).visible_sheet_names


def is_password_protected(file_path: str) -> bool:
    return validate_excel_file(file_path).is_password_protected
