from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_ingest.validation import _resolve_local_path


class FileStatus(Enum):
    VALID               = "VALID"
    NO_HEADERS          = "NO_HEADERS"
    NO_DATA             = "NO_DATA"
    EMPTY_FILE          = "EMPTY_FILE"
    INVALID_STRUCTURE   = "INVALID_STRUCTURE"
    SHEET_NOT_SPECIFIED = "SHEET_NOT_SPECIFIED"

    @property
    def description(self) -> str:
        return {
            "VALID":               "File is readable and has both headers and data rows.",
            "NO_HEADERS":          "No header row detected. File may be raw data — set data_start_row=1 in FileProcessingConfig if intentional.",
            "NO_DATA":             "Header row(s) found but no data rows follow. The sheet may be a template or empty submission form.",
            "EMPTY_FILE":          "Sheet has no rows or columns. Check the file is not corrupt and the correct sheet is selected.",
            "INVALID_STRUCTURE":   "Sheet structure could not be resolved. Review merged cells or irregular layout.",
            "SHEET_NOT_SPECIFIED": "File has multiple visible sheets but no sheet_name was given. Set sheet_name in FileProcessingConfig.",
        }[self.value]

    @property
    def is_actionable(self) -> bool:
        """True when the status requires the caller to take action before proceeding."""
        return self in (
            FileStatus.EMPTY_FILE,
            FileStatus.INVALID_STRUCTURE,
            FileStatus.SHEET_NOT_SPECIFIED,
        )


@dataclass
class MergedCellInfo:
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    span_rows: int
    span_cols: int
    top_left_value: Optional[str]

    @property
    def col_letter_range(self) -> str:
        return f"{get_column_letter(self.min_col)}:{get_column_letter(self.max_col)}"


@dataclass
class HeaderStructure:
    header_row_indices: List[int]          # 1-based row numbers
    num_header_rows: int
    data_start_row: int                    # 1-based
    columns_per_row: Dict[int, int]        # {row_idx: column_count}
    raw_headers: Dict[int, List[Optional[str]]]  # {row_idx: [cell values]}


@dataclass
class FileStructureMetadata:
    sheet_name: str
    status: FileStatus
    header_structure: Optional[HeaderStructure]
    merged_cells: List[MergedCellInfo]
    blank_column_indices: List[int]        # 1-based
    hidden_column_indices: List[int]       # 1-based
    total_rows: int
    total_cols: int
    data_row_count: int
    messages: List[str] = field(default_factory=list)

    @property
    def header_range(self) -> Optional[str]:
        """Excel A1-notation range covering all header rows, e.g. 'A1:L2'. None if no headers."""
        if not self.header_structure or not self.header_structure.header_row_indices:
            return None
        last_col = get_column_letter(self.total_cols) if self.total_cols else "A"
        first_row = self.header_structure.header_row_indices[0]
        last_row  = self.header_structure.header_row_indices[-1]
        return f"A{first_row}:{last_col}{last_row}"

    @property
    def data_range(self) -> Optional[str]:
        """Excel A1-notation range covering all data rows, e.g. 'A2:L21'. None if no data."""
        if not self.header_structure or self.data_row_count == 0:
            return None
        last_col  = get_column_letter(self.total_cols) if self.total_cols else "A"
        first_row = self.header_structure.data_start_row
        last_row  = self.total_rows
        return f"A{first_row}:{last_col}{last_row}"

    def summary_record(self, file_path: str = "", label: str = "") -> dict:
        """Flat dict for Spark DataFrame display — one row per file/sheet.

        Pass ``file_path`` and ``label`` from the caller's FILE_CONFIGS entry so the
        DataFrame row is self-contained without needing a join back to the config list::

            records = [
                framework.detect_structure(path, config=cfg["config"])
                          .summary_record(file_path=cfg["file"], label=cfg["label"])
                for cfg in FILE_CONFIGS
            ]
            display(spark.createDataFrame(records))
        """
        return {
            "label":              label,
            "file":               file_path,
            "sheet_name":         self.sheet_name,
            "status":             self.status.value,
            "status_description": self.status.description,
            "is_actionable":      self.status.is_actionable,
            "total_rows":         self.total_rows,
            "total_cols":         self.total_cols,
            "data_row_count":     self.data_row_count,
            "header_rows":        str(self.header_structure.header_row_indices if self.header_structure else []),
            "header_range":       self.header_range or "",
            "data_range":         self.data_range or "",
            "merged_regions":     len(self.merged_cells),
            "blank_columns":      str(self.blank_column_indices) if self.blank_column_indices else "",
            "hidden_columns":     str(self.hidden_column_indices) if self.hidden_column_indices else "",
        }


@dataclass
class FileProcessingConfig:
    sheet_name: Optional[str] = None
    static_header_rows: Optional[List[int]] = None   # 1-based; if known, skip auto-detect
    data_start_row: Optional[int] = None             # 1-based; if known, skip auto-detect
    ignore_rows: Optional[List[int]] = None          # 1-based rows to skip
    ignore_row_ranges: Optional[List[Tuple[int, int]]] = None
    ignore_columns: Optional[List[int]] = None       # 1-based
    max_rows_to_scan: int = 20
    header_population_threshold: float = 0.3


# ---------------------------------------------------------------------------
# Internal detection helpers
# ---------------------------------------------------------------------------

def _detect_merged_cells(worksheet) -> List[MergedCellInfo]:
    result = []
    for merge in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merge.min_row, merge.min_col).value
        result.append(
            MergedCellInfo(
                min_row=merge.min_row, max_row=merge.max_row,
                min_col=merge.min_col, max_col=merge.max_col,
                span_rows=merge.max_row - merge.min_row + 1,
                span_cols=merge.max_col - merge.min_col + 1,
                top_left_value=str(top_left).strip() if top_left is not None else None,
            )
        )
    return result


def _detect_blank_and_hidden_columns(worksheet) -> Tuple[List[int], List[int]]:
    blank_cols: List[int] = []
    hidden_cols: List[int] = []
    max_col = worksheet.max_column or 0
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = worksheet.column_dimensions.get(letter)
        if dim and dim.hidden:
            hidden_cols.append(col_idx)
        col_values = [
            worksheet.cell(row, col_idx).value
            for row in range(1, min(worksheet.max_row or 1, 20) + 1)
        ]
        if all(v is None or str(v).strip() == "" for v in col_values):
            blank_cols.append(col_idx)
    return blank_cols, hidden_cols


def _row_is_data(row_values: list) -> bool:
    """Return True when the row looks like data rather than a header label row.

    A row is treated as data when more than 35 % of its non-empty cells are
    numeric, date/time, or match a structured-ID pattern (e.g. ORD-xxxx,
    CUST-1234).  Pure-text label rows (headers) score 0 % on this test.
    """
    import re, datetime
    non_empty = [v for v in row_values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return False
    _ID_RE = re.compile(r"^[A-Z]{2,6}-[A-Z0-9]{2,}", re.IGNORECASE)
    data_like = sum(
        1 for v in non_empty
        if isinstance(v, (int, float))
        or isinstance(v, (datetime.datetime, datetime.date))
        or (isinstance(v, str) and _ID_RE.match(v.strip()))
    )
    return (data_like / len(non_empty)) > 0.35


def _detect_header_rows(
    worksheet,
    merged_cells: List[MergedCellInfo],
    config: FileProcessingConfig,
) -> Tuple[HeaderStructure, List[str]]:
    """Returns (HeaderStructure, warning_messages).

    Detection strategy (in priority order):
    1. static_header_rows supplied → trust it, skip auto-detection.
    2. data_start_row == 1 → no header rows (pure data sheet).
    3. Auto-detect: scan rows until the first data-like row (type-based) or
       the first blank gap after header rows are found (population-based).
       If detection hits max_rows_to_scan without finding a clear boundary,
       add a warning and fall back to row 1 as the single header row.
    """
    max_col = worksheet.max_column or 1
    header_rows: List[int] = []
    warnings: List[str] = []

    if config.static_header_rows:
        header_rows = sorted(config.static_header_rows)

    elif config.data_start_row == 1:
        header_rows = []  # caller explicitly said data starts at row 1 — no headers

    else:
        scan_limit = min(config.max_rows_to_scan, worksheet.max_row or 1)
        hit_scan_limit = False
        for row_idx in range(1, scan_limit + 1):
            row_values = [worksheet.cell(row_idx, c).value for c in range(1, max_col + 1)]
            non_empty_count = sum(1 for v in row_values if v is not None and str(v).strip() != "")
            population = non_empty_count / max_col if max_col else 0

            if population < config.header_population_threshold:
                if header_rows:
                    break  # blank gap after header zone — stop
                continue   # leading blank row before headers

            if _row_is_data(row_values):
                break  # first data-like row = end of header zone

            header_rows.append(row_idx)
            if row_idx == scan_limit:
                hit_scan_limit = True

        if hit_scan_limit:
            # Could not find a clear header/data boundary — fall back and warn
            warnings.append(
                f"Header auto-detection reached the scan limit ({scan_limit} rows) without "
                f"finding a clear data boundary. Defaulting to row 1 as the only header row. "
                f"If this is wrong, set static_header_rows in FileProcessingConfig."
            )
            header_rows = [1]

    if not header_rows and config.data_start_row != 1:
        header_rows = [1]

    data_start = config.data_start_row or (max(header_rows) + 1 if header_rows else 1)
    raw_headers: Dict[int, List[Optional[str]]] = {}
    columns_per_row: Dict[int, int] = {}
    for r in header_rows:
        row_vals = [worksheet.cell(r, c).value for c in range(1, max_col + 1)]
        raw_headers[r] = [str(v).strip() if v is not None else None for v in row_vals]
        columns_per_row[r] = max_col

    return HeaderStructure(
        header_row_indices=header_rows,
        num_header_rows=len(header_rows),
        data_start_row=data_start,
        columns_per_row=columns_per_row,
        raw_headers=raw_headers,
    ), warnings


def _count_visible_sheets(workbook) -> int:
    return sum(1 for ws in workbook.worksheets if ws.sheet_state == "visible")


def _resolve_sheet(workbook, config: FileProcessingConfig) -> Tuple[Optional[str], List[str]]:
    msgs: List[str] = []
    if config.sheet_name:
        if config.sheet_name in workbook.sheetnames:
            return config.sheet_name, msgs
        msgs.append(f"Sheet '{config.sheet_name}' not found. Available: {workbook.sheetnames}")
        return None, msgs
    visible_count = _count_visible_sheets(workbook)
    if visible_count == 1:
        sheet = next(ws for ws in workbook.worksheets if ws.sheet_state == "visible")
        msgs.append(f"Single visible sheet auto-selected: '{sheet.title}'")
        return sheet.title, msgs
    msgs.append(
        f"Multiple visible sheets ({visible_count}) — provide sheet_name in FileProcessingConfig."
    )
    return None, msgs


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def analyze_excel_structure(
    file_path: str,
    config: Optional[FileProcessingConfig] = None,
    password: Optional[str] = None,
) -> FileStructureMetadata:
    config = config or FileProcessingConfig()
    local_path = _resolve_local_path(file_path)
    msgs: List[str] = []

    if password is not None:
        try:
            import io, msoffcrypto
        except ImportError:
            raise ImportError(
                "msoffcrypto-tool is required to open AES-encrypted Excel files. "
                "Run: pip install msoffcrypto-tool"
            )
        _use_kwarg = False
        try:
            with open(local_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
            decrypted.seek(0)
            wb = load_workbook(decrypted, read_only=False, data_only=True)
        except Exception as exc:
            msg = str(exc).lower()
            if "password" in msg or "decrypt" in msg:
                raise
            _use_kwarg = True  # not AES-encrypted — fall through to openpyxl worksheet kwarg
        if _use_kwarg:
            wb = load_workbook(local_path, read_only=False, data_only=True, password=password)
    else:
        wb = load_workbook(local_path, read_only=False, data_only=True)

    sheet_name, sheet_msgs = _resolve_sheet(wb, config)
    msgs.extend(sheet_msgs)

    if sheet_name is None:
        wb.close()
        return FileStructureMetadata(
            sheet_name="",
            status=FileStatus.SHEET_NOT_SPECIFIED,
            header_structure=None,
            merged_cells=[], blank_column_indices=[], hidden_column_indices=[],
            total_rows=0, total_cols=0, data_row_count=0, messages=msgs,
        )

    ws = wb[sheet_name]
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0

    if total_rows == 0 or total_cols == 0:
        wb.close()
        return FileStructureMetadata(
            sheet_name=sheet_name, status=FileStatus.EMPTY_FILE,
            header_structure=None,
            merged_cells=[], blank_column_indices=[], hidden_column_indices=[],
            total_rows=total_rows, total_cols=total_cols, data_row_count=0, messages=msgs,
        )

    merged = _detect_merged_cells(ws)
    blank_cols, hidden_cols = _detect_blank_and_hidden_columns(ws)
    header_struct, detection_warnings = _detect_header_rows(ws, merged, config)
    msgs.extend(detection_warnings)

    data_row_count = max(0, total_rows - header_struct.data_start_row + 1)
    if data_row_count == 0:
        status = FileStatus.NO_DATA
    elif not header_struct.header_row_indices:
        status = FileStatus.NO_HEADERS
    else:
        status = FileStatus.VALID

    msgs.append(
        f"Sheet '{sheet_name}': {header_struct.num_header_rows} header row(s), "
        f"{data_row_count} data row(s), {len(merged)} merged region(s)."
    )
    wb.close()

    return FileStructureMetadata(
        sheet_name=sheet_name,
        status=status,
        header_structure=header_struct,
        merged_cells=merged,
        blank_column_indices=blank_cols,
        hidden_column_indices=hidden_cols,
        total_rows=total_rows,
        total_cols=total_cols,
        data_row_count=data_row_count,
        messages=msgs,
    )
