from __future__ import annotations

import io
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from excel_ingest.metadata import MetadataExtractionResult
from excel_ingest.structure import FileProcessingConfig, FileStructureMetadata
from excel_ingest.validation import _resolve_local_path


@dataclass
class LoadResult:
    """Returned by framework.load() — Excel data rows as a Spark DataFrame.

    result.df always contains:
      - All loadable bronze columns (db_canonical_bronze_column_name) as STRING
      - source_file      : filename (basename only, not full path)
      - source_sheet     : sheet name
      - insert_timestamp : datetime when framework.load() was called
    """
    df: Any                              # pyspark.sql.DataFrame
    metadata: MetadataExtractionResult
    messages: List[str] = field(default_factory=list)

    @property
    def source_file(self) -> str:
        return self.metadata.file_metadata.file_name

    @property
    def source_sheet(self) -> str:
        return self.metadata.file_metadata.sheet_name


def _open_for_read(file_path: str, password: Optional[str]):
    """Open workbook in read-only mode, handling AES-encrypted files."""
    from openpyxl import load_workbook
    local_path = _resolve_local_path(file_path)

    if password is not None:
        try:
            import msoffcrypto
        except ImportError:
            raise ImportError(
                "msoffcrypto-tool is required for password-protected files. "
                "Run: pip install msoffcrypto-tool"
            )
        try:
            with open(local_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
            decrypted.seek(0)
            return load_workbook(decrypted, read_only=True, data_only=True)
        except Exception as exc:
            msg = str(exc).lower()
            if "password" in msg or "decrypt" in msg:
                raise
        return load_workbook(local_path, read_only=True, data_only=True, password=password)
    return load_workbook(local_path, read_only=True, data_only=True)


def load_excel_data(
    file_path: str,
    structure: FileStructureMetadata,
    metadata: MetadataExtractionResult,
    spark,
    password: Optional[str] = None,
    config: Optional[FileProcessingConfig] = None,
    skip_hidden_columns: bool = False,
) -> LoadResult:
    """Read data rows from an Excel sheet and return a Spark DataFrame.

    All data columns land as STRING — cast to correct types in the silver layer.
    Three columns are appended automatically to every row:
      - source_file       : basename of the file path
      - source_sheet      : sheet name as read from the workbook
      - insert_timestamp  : Python datetime at the moment load() is called

    Args:
        file_path:            Path to the Excel file (Volume, DBFS, or local).
        structure:            Output of detect_structure() for this file + sheet.
        metadata:             Output of extract_metadata() for this file + sheet.
        spark:                Active SparkSession.
        password:             Password for AES-encrypted files.
        config:               Optional FileProcessingConfig — used only for
                              ignore_rows / ignore_row_ranges at load time.
                              Sheet selection and header detection are already
                              captured in structure; pass config here only when
                              specific data rows must be skipped during loading.
                              Build from a Delta override table using
                              FileProcessingConfig.from_override(row).
        skip_hidden_columns:  If True, columns flagged is_hidden_column are
                              excluded from result.df. Default False — hidden
                              columns carry real data; caller decides whether
                              to include or exclude them.
    """
    msgs: List[str] = []

    # Build column map: col_index → bronze_name (skip blank always; hidden if requested)
    col_map: Dict[int, str] = {}
    for col in metadata.column_metadata:
        if col.is_blank_column:
            continue
        if skip_hidden_columns and col.is_hidden_column:
            continue
        col_map[col.column_index] = col.db_canonical_bronze_column_name

    # Build row-skip set from config overrides
    ignore_set: Set[int] = set()
    if config:
        if config.ignore_rows:
            ignore_set.update(config.ignore_rows)
        if config.ignore_row_ranges:
            for start, end in config.ignore_row_ranges:
                ignore_set.update(range(start, end + 1))

    source_file  = metadata.file_metadata.file_name
    source_sheet = metadata.file_metadata.sheet_name
    load_ts      = datetime.now()

    data_start = (
        structure.header_structure.data_start_row
        if structure.header_structure
        else 1
    )
    total_rows = structure.total_rows

    rows: List[Dict[str, Any]] = []

    if total_rows >= data_start:
        if not col_map:
            msgs.append(
                f"No loadable columns in '{source_sheet}' — all are blank or hidden. "
                f"Check skip_hidden_columns setting."
            )
        else:
            wb = _open_for_read(file_path, password)
            ws = wb[structure.sheet_name]

            for i, row_vals in enumerate(
                ws.iter_rows(min_row=data_start, max_row=total_rows, values_only=True)
            ):
                row_idx = data_start + i
                if row_idx in ignore_set:
                    continue
                row_dict: Dict[str, Any] = {}
                for col_idx, bronze_name in col_map.items():
                    val = row_vals[col_idx - 1] if col_idx - 1 < len(row_vals) else None
                    row_dict[bronze_name] = (
                        str(val).strip()
                        if val is not None and str(val).strip() != ""
                        else None
                    )
                row_dict["source_file"]      = source_file
                row_dict["source_sheet"]     = source_sheet
                row_dict["insert_timestamp"] = load_ts
                rows.append(row_dict)

            wb.close()
            msgs.append(
                f"Loaded {len(rows)} row(s), {len(col_map)} column(s) "
                f"from '{source_sheet}' in '{source_file}'."
            )
    else:
        msgs.append(
            f"No data rows to load from '{source_sheet}' in '{source_file}'."
        )

    # Build Spark DataFrame — always use explicit schema so Spark never needs to infer
    # types. Inference fails when an entire column is NULL (CANNOT_DETERMINE_TYPE).
    # All bronze columns are STRING by design; auto-columns have known fixed types.
    from pyspark.sql.types import StringType, StructField, StructType, TimestampType
    schema = StructType(
        [StructField(name, StringType(), True) for name in col_map.values()]
        + [
            StructField("source_file",      StringType(),    True),
            StructField("source_sheet",     StringType(),    True),
            StructField("insert_timestamp", TimestampType(), True),
        ]
    )
    df = spark.createDataFrame(rows, schema=schema)

    return LoadResult(df=df, metadata=metadata, messages=msgs)


def combine_results(results: List[LoadResult], spark) -> Any:
    """Union multiple LoadResult DataFrames into one, NULL-filling missing columns.

    Computes the full column union across all results. Any bronze column present
    in some files but not others is added as NULL STRING. Auto-columns
    (source_file, source_sheet, insert_timestamp) are preserved at the end of
    every row and are always present.

    Args:
        results: One or more LoadResult from framework.load(). Row order within
                 each result is preserved; results are stacked in list order.
        spark:   Active SparkSession.

    Note:
        No column-overlap check is performed. Combining completely unrelated file
        types (e.g. HR records + invoices) will succeed but produce a very sparse
        table. Caller is responsible for combining only files that share a
        meaningful column set. Document your expected schema in config or a README
        to guard against accidental misuse.
    """
    if not results:
        raise ValueError("combine_results() requires at least one LoadResult.")

    from pyspark.sql.functions import lit
    from pyspark.sql.types import StringType

    AUTO_COLS = ["source_file", "source_sheet", "insert_timestamp"]

    # Sorted union of all bronze column names for a stable, predictable column order
    all_bronze = sorted({
        col.db_canonical_bronze_column_name
        for r in results
        for col in r.metadata.column_metadata
        if not col.is_blank_column
    })
    full_order = all_bronze + AUTO_COLS

    aligned = []
    for result in results:
        df = result.df
        existing = set(df.columns)
        for col_name in all_bronze:
            if col_name not in existing:
                df = df.withColumn(col_name, lit(None).cast(StringType()))
        aligned.append(df.select(full_order))

    combined = aligned[0]
    for df in aligned[1:]:
        combined = combined.union(df)

    return combined
